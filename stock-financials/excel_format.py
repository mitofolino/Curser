"""Excel number formats aligned with header units."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from numbers_format import infer_column_format

_EXCEL_FORMATS = {
    "number": "#,##0.00",
    "number_4": "#,##0.0000",
    "number_6": "0.000000",
    "currency_eur": '#,##0.00 "EUR"',
    "percentage": "0.00%",
    "datetime": "yyyy-mm-dd hh:mm:ss",
}


def _excel_number_format(header: str) -> str | None:
    fmt = infer_column_format(header)
    if fmt.kind == "text":
        return None
    if fmt.kind == "datetime":
        return _EXCEL_FORMATS["datetime"]
    if fmt.kind == "percentage":
        return _EXCEL_FORMATS["percentage"]
    if fmt.kind == "currency" and fmt.currency_code == "EUR":
        return _EXCEL_FORMATS["currency_eur"]
    if fmt.kind == "number":
        if fmt.decimal_places >= 6:
            return _EXCEL_FORMATS["number_6"]
        if fmt.decimal_places >= 4:
            return _EXCEL_FORMATS["number_4"]
        return _EXCEL_FORMATS["number"]
    return None


def apply_sheet_formats(ws: Worksheet, *, data_start_row: int = 2) -> None:
    """Apply number formats to data rows from row-1 headers."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue
        number_format = _excel_number_format(str(header))
        if not number_format:
            continue
        for row in range(data_start_row, ws.max_row + 1):
            cell = ws.cell(row, col_idx)
            if cell.value is not None and cell.value != "":
                cell.number_format = number_format
